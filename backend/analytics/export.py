"""
GET /api/analytics/export/?format=xlsx
GET /api/analytics/export/?format=pdf
"""
import io
from datetime import date

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from accounts.permissions import IsDirector

from .models import AttritionPrediction, EmployeeCluster, Anomaly, MetricForecast


# ── helpers ──────────────────────────────────────────────────────────────────

RISK_RU    = {'high': 'Высокий', 'medium': 'Средний', 'low': 'Низкий'}
SEVERITY_RU = {'high': 'Высокая', 'medium': 'Средняя', 'low': 'Низкая'}
METRIC_RU  = {
    'headcount': 'Численность',
    'turnover':  'Текучесть %',
    'avg_salary': 'Средняя зарплата',
    'sick_days': 'Больничные дни',
    'overtime':  'Сверхурочные часы',
}


def _factors_str(factors):
    if not factors:
        return '—'
    parts = []
    for f in factors[:3]:
        feat  = f.get('label') or f.get('feature', '')
        arrow = '↑' if f.get('direction') == 'up' else '↓'
        parts.append(f"{arrow}{feat}")
    return ', '.join(parts)


# ── Excel ─────────────────────────────────────────────────────────────────────

def _build_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Palette ──────────────────────────────────────────────────────────────
    COLOR_HEADER  = '4F46E5'   # indigo
    COLOR_HIGH    = 'FEE2E2'
    COLOR_MEDIUM  = 'FEF3C7'
    COLOR_LOW     = 'D1FAE5'
    COLOR_GRAY_H  = 'F3F4F6'

    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_font():
        return Font(name='Calibri', bold=True, color='FFFFFF', size=11)

    def header_fill():
        return PatternFill('solid', fgColor=COLOR_HEADER)

    def cell_font():
        return Font(name='Calibri', size=10)

    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    def apply_header(ws, headers, widths):
        ws.row_dimensions[1].height = 22
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font   = header_font()
            cell.fill   = header_fill()
            cell.border = border
            cell.alignment = center()
            ws.column_dimensions[get_column_letter(col)].width = w

    def style_data_row(ws, row, fill_color=None):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.font   = cell_font()
            c.border = border
            c.alignment = left()
            if fill_color:
                c.fill = PatternFill('solid', fgColor=fill_color)

    today_str = date.today().strftime('%d.%m.%Y')

    # ── Sheet 1: Attrition ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Риск увольнения'
    ws1['A1']  # ensure sheet is active
    headers1 = ['ФИО', 'Отдел', 'Должность', 'Оклад, ₽', 'Риск (%)', 'Уровень риска', 'Ключевые факторы', 'Дата прогноза']
    widths1   = [28, 22, 22, 14, 12, 16, 45, 16]
    apply_header(ws1, headers1, widths1)

    preds = AttritionPrediction.objects.select_related(
        'employee', 'employee__department', 'employee__position'
    ).order_by('-risk_score')

    for r, p in enumerate(preds, 2):
        emp = p.employee
        risk_pct = round(p.risk_score * 100, 1)
        ws1.cell(r, 1, emp.full_name)
        ws1.cell(r, 2, emp.department.name if emp.department else '—')
        ws1.cell(r, 3, emp.position.title  if emp.position  else '—')
        ws1.cell(r, 4, float(emp.salary))
        ws1.cell(r, 5, risk_pct)
        ws1.cell(r, 6, RISK_RU.get(p.risk_label, p.risk_label))
        ws1.cell(r, 7, _factors_str(p.top_factors))
        ws1.cell(r, 8, p.predicted_at.strftime('%d.%m.%Y'))
        color = COLOR_HIGH if p.risk_label == 'high' else COLOR_MEDIUM if p.risk_label == 'medium' else COLOR_LOW
        style_data_row(ws1, r, color)

    ws1.freeze_panes = 'A2'

    # ── Sheet 2: Clusters ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Кластеры')
    headers2 = ['ФИО', 'Отдел', 'Кластер №', 'Название кластера', 'X (t-SNE)', 'Y (t-SNE)']
    widths2   = [28, 22, 12, 28, 14, 14]
    apply_header(ws2, headers2, widths2)

    clusters = EmployeeCluster.objects.select_related(
        'employee', 'employee__department'
    ).order_by('cluster_id')

    for r, ec in enumerate(clusters, 2):
        emp = ec.employee
        ws2.cell(r, 1, emp.full_name)
        ws2.cell(r, 2, emp.department.name if emp.department else '—')
        ws2.cell(r, 3, ec.cluster_id)
        ws2.cell(r, 4, ec.cluster_label or '—')
        ws2.cell(r, 5, round(ec.x_tsne, 4))
        ws2.cell(r, 6, round(ec.y_tsne, 4))
        fill = COLOR_GRAY_H if r % 2 == 0 else 'FFFFFF'
        style_data_row(ws2, r, fill)

    ws2.freeze_panes = 'A2'

    # ── Sheet 3: Anomalies ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Аномалии')
    headers3 = ['Сотрудник', 'Метрика', 'Значение', 'Ожидаемое', 'Anomaly Score', 'Серьёзность', 'Описание', 'Дата', 'Решено']
    widths3   = [28, 22, 12, 12, 15, 14, 50, 14, 10]
    apply_header(ws3, headers3, widths3)

    anomalies = Anomaly.objects.select_related('employee').order_by('-detected_at')

    for r, a in enumerate(anomalies, 2):
        ws3.cell(r, 1, a.employee.full_name if a.employee else 'Системная')
        ws3.cell(r, 2, a.metric)
        ws3.cell(r, 3, round(a.value, 2))
        ws3.cell(r, 4, round(a.expected_value, 2) if a.expected_value is not None else '—')
        ws3.cell(r, 5, round(a.anomaly_score, 4))
        ws3.cell(r, 6, SEVERITY_RU.get(a.severity, a.severity))
        ws3.cell(r, 7, a.description)
        ws3.cell(r, 8, a.detected_at.strftime('%d.%m.%Y'))
        ws3.cell(r, 9, 'Да' if a.is_resolved else 'Нет')
        color = COLOR_HIGH if a.severity == 'high' else COLOR_MEDIUM if a.severity == 'medium' else COLOR_GRAY_H
        style_data_row(ws3, r, color)

    ws3.freeze_panes = 'A2'

    # ── Sheet 4: Forecasts ────────────────────────────────────────────────────
    ws4 = wb.create_sheet('Прогнозы SARIMA')
    headers4 = ['Метрика', 'Период', 'Прогноз', 'Нижняя граница', 'Верхняя граница']
    widths4   = [26, 14, 16, 18, 18]
    apply_header(ws4, headers4, widths4)

    forecasts = MetricForecast.objects.order_by('metric', 'period')

    for r, f in enumerate(forecasts, 2):
        ws4.cell(r, 1, METRIC_RU.get(f.metric, f.metric))
        ws4.cell(r, 2, f.period.strftime('%m.%Y'))
        ws4.cell(r, 3, round(f.forecast_value, 2))
        ws4.cell(r, 4, round(f.lower_bound, 2) if f.lower_bound is not None else '—')
        ws4.cell(r, 5, round(f.upper_bound, 2) if f.upper_bound is not None else '—')
        fill = COLOR_GRAY_H if r % 2 == 0 else 'FFFFFF'
        style_data_row(ws4, r, fill)

    ws4.freeze_panes = 'A2'

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws0 = wb.create_sheet('Сводка', 0)
    ws0.column_dimensions['A'].width = 36
    ws0.column_dimensions['B'].width = 20

    title_font = Font(name='Calibri', bold=True, size=16, color=COLOR_HEADER)
    sub_font   = Font(name='Calibri', bold=True, size=12)
    val_font   = Font(name='Calibri', size=11)

    ws0['A1'] = 'Отчёт HR-аналитики'
    ws0['A1'].font = title_font
    ws0['A2'] = f'Дата формирования: {today_str}'
    ws0['A2'].font = val_font

    rows = [
        ('', ''),
        ('Показатель', 'Значение'),
        ('Сотрудников в прогнозе', preds.count()),
        ('  — Высокий риск',  preds.filter(risk_label='high').count()),
        ('  — Средний риск',  preds.filter(risk_label='medium').count()),
        ('  — Низкий риск',   preds.filter(risk_label='low').count()),
        ('', ''),
        ('Кластеров выявлено', EmployeeCluster.objects.values('cluster_id').distinct().count()),
        ('Сотрудников в кластерах', clusters.count()),
        ('', ''),
        ('Аномалий всего', anomalies.count()),
        ('  — Нерешённых', anomalies.filter(is_resolved=False).count()),
        ('  — Высокой серьёзности', anomalies.filter(severity='high').count()),
        ('', ''),
        ('Прогнозных точек SARIMA', forecasts.count()),
    ]

    for i, (label, value) in enumerate(rows, 4):
        c_label = ws0.cell(i, 1, label)
        c_val   = ws0.cell(i, 2, value)
        if label == 'Показатель':
            c_label.font = sub_font
            c_val.font   = sub_font
            c_label.fill = PatternFill('solid', fgColor=COLOR_HEADER)
            c_val.fill   = PatternFill('solid', fgColor=COLOR_HEADER)
            c_label.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
            c_val.font   = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        else:
            c_label.font = val_font
            c_val.font   = val_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF ───────────────────────────────────────────────────────────────────────

def _build_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # Register a Unicode-capable font if available, fall back to Helvetica
    FONT_NAME = 'Helvetica'
    try:
        # Try DejaVu from system or common locations
        candidates = [
            'C:/Windows/Fonts/arial.ttf',
            'C:/Windows/Fonts/calibri.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        ]
        for path in candidates:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont('UniFont', path))
                FONT_NAME = 'UniFont'
                break
    except Exception:
        pass

    today_str = date.today().strftime('%d.%m.%Y')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    INDIGO  = colors.HexColor('#4F46E5')
    GRAY_BG = colors.HexColor('#F3F4F6')
    RED_BG  = colors.HexColor('#FEE2E2')
    YEL_BG  = colors.HexColor('#FEF3C7')
    GRN_BG  = colors.HexColor('#D1FAE5')

    title_style = ParagraphStyle('Title', fontName=FONT_NAME, fontSize=18,
                                  textColor=INDIGO, spaceAfter=4, leading=22)
    sub_style   = ParagraphStyle('Sub',   fontName=FONT_NAME, fontSize=10,
                                  textColor=colors.gray, spaceAfter=12)
    section_style = ParagraphStyle('Section', fontName=FONT_NAME, fontSize=13,
                                    textColor=INDIGO, spaceBefore=14, spaceAfter=6,
                                    borderPad=4)

    def _tbl_style(header_bg=INDIGO, stripe=GRAY_BG):
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), FONT_NAME),
            ('FONTSIZE',   (0, 0), (-1, 0), 9),
            ('FONTNAME',   (0, 1), (-1, -1), FONT_NAME),
            ('FONTSIZE',   (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, stripe]),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#D1D5DB')),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ])

    story = []

    # ── Title ─────────────────────────────────────────────────────────────────
    story.append(Paragraph('Отчёт HR-аналитики', title_style))
    story.append(Paragraph(f'Дата формирования: {today_str}', sub_style))
    story.append(HRFlowable(width='100%', color=INDIGO, thickness=1))
    story.append(Spacer(1, 0.3*cm))

    W = landscape(A4)[0] - 3*cm  # usable width

    # ── Summary ───────────────────────────────────────────────────────────────
    preds     = AttritionPrediction.objects.select_related('employee', 'employee__department')
    anomalies = Anomaly.objects.select_related('employee')
    clusters  = EmployeeCluster.objects.select_related('employee', 'employee__department')
    forecasts = MetricForecast.objects.order_by('metric', 'period')

    summary_data = [
        ['Показатель', 'Значение'],
        ['Сотрудников в прогнозе', str(preds.count())],
        ['  Высокий риск',         str(preds.filter(risk_label='high').count())],
        ['  Средний риск',         str(preds.filter(risk_label='medium').count())],
        ['  Низкий риск',          str(preds.filter(risk_label='low').count())],
        ['Аномалий (нерешённых)',  str(anomalies.filter(is_resolved=False).count())],
        ['Прогнозных точек SARIMA', str(forecasts.count())],
    ]
    t = Table(summary_data, colWidths=[W * 0.55, W * 0.2])
    t.setStyle(_tbl_style())
    story.append(t)
    story.append(Spacer(1, 0.4*cm))

    # ── Attrition ─────────────────────────────────────────────────────────────
    story.append(Paragraph('Прогноз риска увольнения (XGBoost)', section_style))
    header = ['ФИО сотрудника', 'Отдел', 'Риск (%)', 'Уровень', 'Ключевые факторы']
    cw = [W*0.25, W*0.18, W*0.09, W*0.10, W*0.38]
    rows = [header]
    for p in preds.order_by('-risk_score')[:30]:
        emp = p.employee
        rows.append([
            emp.full_name,
            emp.department.name if emp.department else '—',
            f'{p.risk_score*100:.1f}%',
            RISK_RU.get(p.risk_label, p.risk_label),
            _factors_str(p.top_factors),
        ])

    ts = _tbl_style()
    # Colour high/medium/low rows
    for i, p in enumerate(preds.order_by('-risk_score')[:30], 1):
        bg = RED_BG if p.risk_label == 'high' else YEL_BG if p.risk_label == 'medium' else GRN_BG
        ts.add('BACKGROUND', (0, i), (-1, i), bg)

    t = Table(rows, colWidths=cw)
    t.setStyle(ts)
    story.append(t)
    story.append(Spacer(1, 0.4*cm))

    # ── Anomalies ─────────────────────────────────────────────────────────────
    story.append(Paragraph('Обнаруженные аномалии (Isolation Forest)', section_style))
    header = ['Сотрудник', 'Метрика', 'Значение', 'Ожидаемое', 'Score', 'Серьёзность', 'Описание', 'Дата']
    cw2 = [W*0.16, W*0.12, W*0.08, W*0.08, W*0.08, W*0.10, W*0.24, W*0.10]
    rows2 = [header]
    for a in anomalies.order_by('-detected_at')[:25]:
        rows2.append([
            a.employee.full_name if a.employee else 'Системная',
            a.metric,
            str(round(a.value, 2)),
            str(round(a.expected_value, 2)) if a.expected_value is not None else '—',
            str(round(a.anomaly_score, 3)),
            SEVERITY_RU.get(a.severity, a.severity),
            a.description[:80],
            a.detected_at.strftime('%d.%m.%Y'),
        ])
    t2 = Table(rows2, colWidths=cw2)
    t2.setStyle(_tbl_style())
    story.append(t2)
    story.append(Spacer(1, 0.4*cm))

    # ── Forecasts ─────────────────────────────────────────────────────────────
    story.append(Paragraph('Прогнозы HR-метрик (SARIMA, 95% ДИ)', section_style))
    header3 = ['Метрика', 'Период', 'Прогноз', 'Нижняя граница', 'Верхняя граница']
    cw3 = [W*0.25, W*0.12, W*0.15, W*0.18, W*0.18]
    rows3 = [header3]
    for f in forecasts:
        rows3.append([
            METRIC_RU.get(f.metric, f.metric),
            f.period.strftime('%m.%Y'),
            str(round(f.forecast_value, 2)),
            str(round(f.lower_bound, 2)) if f.lower_bound is not None else '—',
            str(round(f.upper_bound, 2)) if f.upper_bound is not None else '—',
        ])
    t3 = Table(rows3, colWidths=cw3)
    t3.setStyle(_tbl_style())
    story.append(t3)

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width='100%', color=colors.lightgrey, thickness=0.5))
    story.append(Paragraph(
        f'Сформировано системой HRM Analytics · {today_str}',
        ParagraphStyle('footer', fontName=FONT_NAME, fontSize=8, textColor=colors.gray),
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ── View ──────────────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsDirector])
def export_analytics(request):
    from rest_framework.response import Response as DRFResponse
    import traceback

    # NOTE: 'format' is reserved by DRF for content negotiation — use 'export_format'
    fmt = request.query_params.get('export_format', 'xlsx').lower()
    today = date.today().strftime('%Y-%m-%d')

    if fmt not in ('xlsx', 'pdf'):
        return DRFResponse(
            {'detail': 'Неизвестный формат. Используйте xlsx или pdf.'},
            status=400,
        )

    try:
        if fmt == 'xlsx':
            buf = _build_xlsx()
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="hr_analytics_{today}.xlsx"'
            return response

        buf = _build_pdf()
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="hr_analytics_{today}.pdf"'
        return response

    except Exception as exc:
        traceback.print_exc()
        return DRFResponse(
            {'detail': f'Ошибка генерации {fmt.upper()}: {exc}'},
            status=500,
        )
